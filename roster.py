import datetime
import pandas as pd
import random
from collections import defaultdict

ALL_DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
WEEKENDS = ["Saturday", "Sunday"]

class Staff:
    def __init__(self, name, role, availability, max_hours=9999, min_off_days=2):
        self.name = name
        self.role = role
        self.availability = availability
        self.max_hours = max_hours
        self.min_off_days = min_off_days
        self.weekly_off_requests = {}
        self.schedule = {}
        self.total_hours = 0

    def assign_shift(self, day, start, end, label=None):
        duration = datetime.datetime.combine(datetime.date.today(), end) - datetime.datetime.combine(datetime.date.today(), start)
        hours = duration.total_seconds() / 3600
        self.schedule[day] = (start, end)
        self.total_hours += hours

    def is_available(self, day):
        return day in self.availability and self.schedule.get(day) is None

class RosterGenerator:
    def __init__(self, staff_list,
                 opening_hour="11:00", closing_hour="21:00",
                 report_lead=30,
                 min_staff_weekday=6, min_staff_weekend=6,
                 training_schedule=None, activities=None,
                 enforce_non_consecutive_closing=True,
                 enforce_non_consecutive_incharge=True,
                 max_closing_per_week=None,
                 max_incharge_per_week=None,
                 auto_tune_enabled=True):

        self.staff_list = staff_list
        self.opening_time = datetime.datetime.strptime(opening_hour, "%H:%M").time()
        self.closing_time = datetime.datetime.strptime(closing_hour, "%H:%M").time()
        self.report_time = self._subtract_minutes(self.opening_time, report_lead)
        self.morning_end = datetime.time(19, 0)
        self.afternoon_start = datetime.time(12, 0)
        self.min_weekday = min_staff_weekday
        self.min_weekend = min_staff_weekend
        self.training_schedule = training_schedule or {}
        self.activities = activities or {}
        self.enforce_non_consecutive_closing = enforce_non_consecutive_closing
        self.enforce_non_consecutive_incharge = enforce_non_consecutive_incharge
        self.max_closing_per_week = max_closing_per_week
        self.max_incharge_per_week = max_incharge_per_week
        self.auto_tune_enabled = auto_tune_enabled
        self.roster = pd.DataFrame(index=[s.name for s in staff_list], columns=ALL_DAYS)
        self.violations = []

    def _subtract_minutes(self, time_obj, minutes):
        return (datetime.datetime.combine(datetime.date.today(), time_obj) - datetime.timedelta(minutes=minutes)).time()

    def assign_off_days(self, week_id):
        for staff in self.staff_list:
            requested = staff.weekly_off_requests.get(week_id, [])
            scheduled = list(staff.schedule.keys())
            selected = [d for d in requested if d in staff.availability and d not in scheduled]

            remaining = staff.min_off_days - len(selected)
            fillable = [d for d in ALL_DAYS if d in staff.availability and d not in scheduled and d not in selected]
            random.shuffle(fillable)
            selected += fillable[:remaining]

            for day in selected:
                staff.schedule[day] = None
                self.roster.at[staff.name, day] = "OFF"

    def assign_daily_in_charge(self):
        pool = [s for s in self.staff_list if len(s.availability) >= 5]
        count_tracker = defaultdict(int)

        for i, day in enumerate(ALL_DAYS):
            prev = ALL_DAYS[i - 1] if i > 0 else None
            eligible = []
            for s in pool:
                had_incharge = prev in s.schedule and isinstance(s.schedule[prev], str) and "In-Charge" in s.schedule[prev]
                if self.enforce_non_consecutive_incharge and had_incharge:
                    continue
                if self.max_incharge_per_week and count_tracker[s.name] >= self.max_incharge_per_week:
                    continue
                if s.is_available(day):
                    eligible.append(s)
            if eligible:
                selected = sorted(eligible, key=lambda s: count_tracker[s.name])[0]
                selected.assign_shift(day, datetime.time(10, 0), datetime.time(22, 0), "In-Charge")
                self.roster.at[selected.name, day] = "In-Charge: 10:00–22:00"
                count_tracker[selected.name] += 1
            else:
                self.violations.append(f"No in-charge available on {day}")

    def assign_closing_staff(self):
        count_tracker = defaultdict(int)
        for i, day in enumerate(ALL_DAYS):
            prev = ALL_DAYS[i - 1] if i > 0 else None
            eligible = []
            for s in self.staff_list:
                closed_yesterday = prev in s.schedule and isinstance(s.schedule[prev], tuple) and s.schedule[prev][1] == datetime.time(22, 0)
                if self.enforce_non_consecutive_closing and closed_yesterday:
                    continue
                if self.max_closing_per_week and count_tracker[s.name] >= self.max_closing_per_week:
                    continue
                if s.is_available(day):
                    eligible.append(s)
            if eligible:
                selected = sorted(eligible, key=lambda s: count_tracker[s.name])[0]
                selected.assign_shift(day, self.report_time, datetime.time(22, 0), "Closing")
                self.roster.at[selected.name, day] = f"Closing: {self.report_time.strftime('%H:%M')}–22:00"
                count_tracker[selected.name] += 1
            else:
                self.violations.append(f"No closing staff available on {day}")

    def fill_remaining_shifts(self, day, required_count, shift_tracker):
        current = sum(
            1 for s in self.staff_list
            if self.roster.at[s.name, day] not in ["OFF", None]
        )
        shortfall = max(0, required_count - current)
        eligible = [s for s in self.staff_list if s.is_available(day)]

        # Limit 1 morning and 1 afternoon per day
        morning_given = False
        afternoon_given = False

        for s in sorted(eligible, key=lambda x: x.total_hours):
            name = s.name
            if shift_tracker[name]["Morning"] == 0 and not morning_given:
                s.assign_shift(day, self.report_time, self.morning_end, "Morning")
                self.roster.at[name, day] = f"Morning: {self.report_time.strftime('%H:%M')}–{self.morning_end.strftime('%H:%M')}"
                shift_tracker[name]["Morning"] += 1
                morning_given = True
            elif shift_tracker[name]["Afternoon"] == 0 and not afternoon_given:
                s.assign_shift(day, self.afternoon_start, self.closing_time, "Afternoon")
                self.roster.at[name, day] = f"Afternoon: {self.afternoon_start.strftime('%H:%M')}–{self.closing_time.strftime('%H:%M')}"
                shift_tracker[name]["Afternoon"] += 1
                afternoon_given = True

            current = sum(
                1 for s in self.staff_list
                if self.roster.at[s.name, day] not in ["OFF", None]
            )
            if current >= required_count:
                break

    def generate(self, week_id="Week 1"):
        shift_tracker = {s.name: {"Morning": 0, "Afternoon": 0} for s in self.staff_list}
        self.assign_off_days(week_id)
        self.assign_daily_in_charge()
        self.assign_closing_staff()

        for day in ALL_DAYS:
            required = self.min_weekend if day in WEEKENDS else self.min_weekday
            self.fill_remaining_shifts(day, required, shift_tracker)

        for staff in self.staff_list:
            for day in ALL_DAYS:
                if day not in staff.schedule:
                    staff.schedule[day] = None
                    self.roster.at[staff.name, day] = "OFF"
                    self.violations.append(f"{staff.name} was auto-marked OFF on {day} due to no assignment")

        return self.roster

    def summary(self):
        return pd.DataFrame({
            "Staff": [s.name for s in self.staff_list],
            "Total Hours": [round(s.total_hours, 2) for s in self.staff_list],
            "Scheduled Days": [len([d for d, v in s.schedule.items() if v]) for s in self.staff_list],
            "Off-Days": [len([d for d, v in s.schedule.items() if v is None]) for s in self.staff_list]
        })

    def list_violations(self):
        return self.violations

    def export_to_excel(self, filename="weekly_roster.xlsx"):
        from openpyxl import Workbook        
        from openpyxl.styles import PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Roster"

        ws.append(["Activity"] + [self.activities.get(day, "—") for day in ALL_DAYS])
        ws.append(["Staff"] + ALL_DAYS)

        for staff in self.staff_list:
            row = [staff.name]
            for day in ALL_DAYS:
                row.append(self.roster.at[staff.name, day] or "")
            ws.append(row)

        fill_colors = {
            "Morning": "ADD8E6",
            "Afternoon": "F5DEB3",
            "In-Charge": "FFA500",
            "Closing": "87CEEB",
            "Training": "90EE90",
            "OFF": "D3D3D3"
        }

        for row in ws.iter_rows(min_row=3, min_col=2):
            for cell in row:
                value = cell.value or ""
                for label, color in fill_colors.items():
                    if label in value or value == label:
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        wb.save(filename)
